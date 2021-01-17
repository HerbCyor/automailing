from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from mailsender import send_email

    
def read_excel(source_path):
    
    '''
        Reads .xlsx spreadsheed in the format [e-mail,name,info1,info2,info3...]
        and returns a list of lists with all the contacts.
    '''

    #loads a .xlsx file onto workbook object
    wb = load_workbook(source_path)
    #sets active tab
    sheet = wb.active
    
    outer_list = []
    #iterate through rows and collumns, getting the values of the rows and keeping them together in a list (inner_list)
    #this code starts from row 2 adn max_row should be the number of contacts in your list.
    for row in sheet.iter_rows(2,max_row=sheet.max_row):
        inner_list = []
        for cell in row:
            if cell.value != None: #this avoids getting blank cells
                inner_list.append(cell.value)
        #appends inner_list of each row to a BIG LIST with all the contacts (outer_list), given it has data
        if len(inner_list) > 0:
            outer_list.append(inner_list)
    
    return outer_list

def write_on_image (template,txt_source):
    '''
        Writes personal data extracted from an .xlsx spreadsheet onto a personalized picture. Meant for creating
        multiple stances of the same picture, thus creating multiple business cards, for example.

        template used: Designed by vectorstock (Image #16322660 at VectorStock.com)
    '''

    with open (template, "rb") as fp:
        
        #Text variables as position, color, size and font. relative path to the font file.
        
        text_position_X = 150
        text_position_Y = 740
        text_color = (235,135,60)
        text_font = 'Lato-Black.ttf'
        font_size = 25
        
        #Initialization of the font and image
        fnt = ImageFont.truetype(text_font, font_size)
        img = Image.open(fp)
        d = ImageDraw.Draw(img)

        txt = ''
        #Concatenates all the information on the cells of a given row into 'txt'
        for entry in txt_source[1:]:
            txt = txt + entry + "\n"

        
        #adds the text to the image
        d.multiline_text((text_position_X,text_position_Y), txt, font=fnt, align='center', fill=text_color)
        output_file = txt_source[0] + '.png' #the format can be changed
        #Saves the picture into the root folder. the name of the picture is the first column of each row.
        img.save(output_file)

def create_and_send(template_name,excel_list):
    '''
        This function executes all the steps of reading .xlsx, creating the images and sending them to contacts via e-mail. 
    '''
    #creates a list of lists from excel

    text_from_excel = read_excel(excel_list)
    
    #iterates through all the inner lists
    for entry in text_from_excel:
        write_on_image(template_name, entry) #creates a specific imagem for each person

        #personalized email message
        email_subject = "subject"
        email_message = "Hello " + entry[1] + ".\nhere goes your e-mail body message." 
        
        attachment = entry[0] + ".png"
        #sends the e-mail
        send_email(entry[0],email_subject,email_message,attachment_location=attachment)


if __name__ == '__main__':
    
    template_name = 'exampletemplate.png'
    excel_list = 'examplelist.xlsx'
    
    create_and_send(template_name,excel_list)